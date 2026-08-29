import io

from openpyxl import load_workbook

from backend.src.services.xlsx_exporter import export_issues_xlsx


def _issue(iid: str, criterion: str, severity: str, level: str = "AA") -> dict:
    return {
        "id": iid,
        "url": "https://example.com/checkout",
        "guideline": "WCAG 2.2",
        "criterion": criterion,
        "criterion_pt": criterion,
        "severity": severity,
        "severity_pt": severity,
        "element": "<button class='btn-pay'>Pagar</button>",
        "description": "Botão de pagamento sem contraste de texto suficiente",
        "description_technical": "color #999 on #fff < 4.5:1",
        "why_simple": "Pessoas com baixa visão não conseguem identificar o botão",
        "why_technical": "Ratio 2.8:1 abaixo do limite AA",
        "suggestion": "Ajuste a cor do texto para #1a1a1a",
        "suggestion_technical": "color: #1a1a1a",
        "level": level,
        "wcag_url": "https://www.w3.org/WAI/WCAG22/Understanding/contrast-minimum",
        "fixed_element_html": "<button class='btn-pay' style='color:#1a1a1a'>Pagar</button>",
    }


def _mixed_issues() -> list[dict]:
    return [
        _issue("iss-4", "4.1.2 Name, Role, Value", "critical", "A"),
        _issue("iss-1", "1.1.1 Non-text Content", "high", "A"),
        _issue("iss-3", "3.1.1 Language of Page", "medium", "A"),
        _issue("iss-2", "2.1.1 Keyboard", "low", "A"),
    ]


class TestExportIssuesXlsx:
    def test_returns_valid_xlsx_bytes(self) -> None:
        data = export_issues_xlsx([_issue("x-1", "1.4.3 Contrast Minimum", "high", "AA")])
        assert isinstance(data, bytes)
        assert len(data) > 0
        # Assinatura de arquivo ZIP/XLSX
        assert data[:2] == b"PK"

    def test_structure_headers_and_rows(self) -> None:
        issues = _mixed_issues()
        wb = load_workbook(io.BytesIO(export_issues_xlsx(issues)))
        ws = wb.active
        assert ws is not None
        assert ws.title == "Issues de Acessibilidade"

        headers = [c.value for c in ws[1]]
        assert headers[0] == "ID"
        assert "Site / URL Testada" in headers
        assert "Onde Está o Problema (Localização / Elemento)" in headers
        assert "Princípio WCAG" in headers
        assert "Diretriz WCAG" in headers
        assert "Critério de Sucesso" in headers
        assert "Criticidade Normativa (WCAG)" in headers
        assert "Prioridade de Atendimento" in headers
        assert "O Que Aconteceu (Diagnóstico)" in headers
        assert "Por Que Afeta o Usuário" in headers
        assert "Como Resolver (Plano de Remediação)" in headers
        assert "Navegador Usado" in headers
        assert "Leitor de Tela Testado" in headers
        assert "Status" in headers
        assert len(headers) == 16

        # 1 header + 4 linhas de dados
        assert ws.max_row == len(issues) + 1
        # freeze panes no header
        assert ws.freeze_panes == "A2"
        # auto-filter aplicado
        assert ws.auto_filter.ref is not None

    def test_sorts_rows_by_wcag_principle_order(self) -> None:
        """Verifica que as linhas são ordenadas pelos princípios POUR de cima para baixo."""
        issues = _mixed_issues()
        wb = load_workbook(io.BytesIO(export_issues_xlsx(issues)))
        ws = wb.active
        assert ws is not None

        # Coluna 4 = Princípio WCAG
        principles = [ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)]
        assert principles[0] == "Princípio 1: Perceptível"
        assert principles[1] == "Princípio 2: Operável"
        assert principles[2] == "Princípio 3: Compreensível"
        assert principles[3] == "Princípio 4: Robusto"

    def test_site_and_what_was_tested_columns_are_populated(self) -> None:
        issues = _mixed_issues()
        wb = load_workbook(io.BytesIO(export_issues_xlsx(issues)))
        ws = wb.active
        assert ws is not None
        headers = [c.value for c in ws[1]]
        site_col = headers.index("Site / URL Testada") + 1
        tested_col = headers.index("Onde Está o Problema (Localização / Elemento)") + 1
        nav_col = headers.index("Navegador Usado") + 1
        sr_col = headers.index("Leitor de Tela Testado") + 1
        status_col = headers.index("Status") + 1

        for row_idx in range(2, ws.max_row + 1):
            assert ws.cell(row=row_idx, column=site_col).value == "https://example.com/checkout"
            assert "<button" in ws.cell(row=row_idx, column=tested_col).value
            assert "Chromium" in ws.cell(row=row_idx, column=nav_col).value
            assert "NVDA" in ws.cell(row=row_idx, column=sr_col).value
            assert ws.cell(row=row_idx, column=status_col).value in ("Aberto", "Corrigido")

    def test_severity_and_pedagogical_columns(self) -> None:
        issues = _mixed_issues()
        wb = load_workbook(io.BytesIO(export_issues_xlsx(issues)))
        ws = wb.active
        assert ws is not None
        headers = [c.value for c in ws[1]]
        crit_col = headers.index("Criticidade Normativa (WCAG)") + 1
        prio_col = headers.index("Prioridade de Atendimento") + 1
        prob_col = headers.index("O Que Aconteceu (Diagnóstico)") + 1
        impact_col = headers.index("Por Que Afeta o Usuário") + 1
        sol_col = headers.index("Como Resolver (Plano de Remediação)") + 1

        for row_idx in range(2, ws.max_row + 1):
            assert ws.cell(row=row_idx, column=crit_col).value in ("Nível A", "Nível AA", "Nível AAA")
            assert "Prioridade" in ws.cell(row=row_idx, column=prio_col).value
            assert ws.cell(row=row_idx, column=prob_col).value
            assert ws.cell(row=row_idx, column=impact_col).value
            assert ws.cell(row=row_idx, column=sol_col).value

    def test_empty_list_produces_workbook(self) -> None:
        data = export_issues_xlsx([])
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws is not None
        assert ws.max_row == 1  # só o header
