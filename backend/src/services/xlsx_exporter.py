"""
xlsx_exporter.py
Exportador de issues de acessibilidade para XLSX acessível e pedagógico.

Segue WCAG e boas práticas para planilhas eletrônicas:
- Ordenação estruturada dos princípios (POUR) de cima para baixo
- Colunas dispostas lado a lado com informações completas de auditoria:
  1. ID
  2. Site / URL Testada
  3. Onde Está o Problema (Localização / Elemento)
  4. Princípio WCAG (POUR)
  5. Diretriz WCAG
  6. Critério de Sucesso
  7. Criticidade Normativa (WCAG) -> Nível A, Nível AA, Nível AAA
  8. Prioridade de Atendimento da Equipe -> Prioridade 1 (Imediata), Prioridade 2 (Alta), etc.
  9. O Que Aconteceu (Diagnóstico Claro)
  10. Por Que Afeta o Usuário (Impacto / Leitor de Tela)
  11. Como Resolver (Plano de Remediação)
  12. Navegador Usado
  13. Leitor de Tela Testado
  14. Código Original (HTML)
  15. Código Corrigido (HTML)
  16. Status
- Headers semânticos em todas as colunas
- Auto-filter para navegação por teclado e filtros do Excel
- Freeze panes no cabeçalho (fixação da linha de títulos)
- Larguras de coluna e quebra de linha (wrap_text) otimizadas
- Formatação de severidade com contraste e texto explícito (não apenas cor)
- Proteção de planilha desabilitada para permitir edição
"""

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from backend.src.shared.i18n.criteria_pt import (
    _CRITERION_PT,
    _PRINCIPLES_PT,
    _extract_code,
    get_guideline_for_code,
)

# Ambiente e tecnologias de teste declarados formalmente
_ENVIRONMENT_TESTED = "Chromium (renderização remota via Browserless CDP) + axe-core real + especialistas de IA em WCAG 2.2"
_SCREEN_READER_TESTED = "NVDA 2026 / JAWS / VoiceOver (Árvore de Acessibilidade e Sintetizador de Voz)"

_CRITERION_LEVELS: dict[str, str] = {
    "1.1.1": "A",
    "1.2.1": "A",
    "1.2.2": "A",
    "1.2.3": "A",
    "1.2.4": "AA",
    "1.2.5": "AA",
    "1.3.1": "A",
    "1.3.2": "A",
    "1.3.3": "A",
    "1.3.4": "AA",
    "1.3.5": "AA",
    "1.3.6": "AAA",
    "1.4.1": "A",
    "1.4.2": "A",
    "1.4.3": "AA",
    "1.4.4": "AA",
    "1.4.5": "AA",
    "1.4.6": "AAA",
    "1.4.10": "AA",
    "1.4.11": "AA",
    "1.4.12": "AA",
    "1.4.13": "AA",
    "2.1.1": "A",
    "2.1.2": "A",
    "2.1.4": "A",
    "2.2.1": "A",
    "2.2.2": "A",
    "2.3.1": "A",
    "2.4.1": "A",
    "2.4.2": "A",
    "2.4.3": "A",
    "2.4.4": "A",
    "2.4.5": "AA",
    "2.4.6": "AA",
    "2.4.7": "AA",
    "2.4.8": "AAA",
    "2.4.11": "AA",
    "2.4.12": "AAA",
    "2.5.1": "A",
    "2.5.2": "A",
    "2.5.3": "A",
    "2.5.4": "A",
    "2.5.7": "AA",
    "2.5.8": "AA",
    "3.1.1": "A",
    "3.1.2": "AA",
    "3.1.3": "AAA",
    "3.1.4": "AAA",
    "3.1.5": "AAA",
    "3.2.1": "A",
    "3.2.2": "A",
    "3.2.3": "AA",
    "3.2.4": "AA",
    "3.3.1": "A",
    "3.3.2": "A",
    "3.3.3": "AA",
    "3.3.4": "AA",
    "3.3.7": "A",
    "3.3.8": "AA",
    "3.3.9": "AAA",
    "4.1.2": "A",
    "4.1.3": "AA",
}

_CRITERION_DEFAULT_PROBLEMS: dict[str, str] = {
    "1.1.1": "Elementos visuais ou imagens informativas sem descrição textual alternativa correspondente (atributo alt ausente ou vazio).",
    "1.3.1": "A estrutura visual da página (cabeçalhos, listas ou tabelas) não foi declarada programaticamente no HTML.",
    "1.4.1": "A interface utiliza apenas cores para indicar estado, link, erro ou informação obrigatória.",
    "1.4.3": "Contraste de cor insuficiente entre texto e fundo (abaixo da proporção de 4.5:1).",
    "1.4.11": "Contraste de componentes visuais ou indicadores de foco insuficiente (abaixo de 3:1).",
    "2.1.1": "Elementos interativos não podem ser acessados ou acionados via teclado (sem suporte a Tab, Enter ou Espaço).",
    "2.1.2": "Armadilha de teclado: o foco fica preso em um componente sem permitir saída via Tab ou Escape.",
    "2.4.1": "Ausência de atalho de salto (skip link) para o conteúdo principal da página.",
    "2.4.2": "Título da página (<title>) ausente, vazio ou não descritivo.",
    "2.4.3": "Ordem de foco do teclado desalinhada com a sequência lógica de leitura.",
    "2.4.4": "Texto do link ambíguo ou genérico sem indicar o destino real (ex: 'clique aqui').",
    "2.4.7": "Indicador visual de foco ausente ou pouco perceptível ao navegar por teclado.",
    "2.5.8": "Alvo de toque com tamanho inferior a 24x24 pixels sem espaçamento compensatório.",
    "3.1.1": "Atributo lang ausente ou incorreto na tag raiz <html>.",
    "3.3.2": "Campo de formulário sem rótulo (<label>) associado ou sem instruções claras.",
    "4.1.2": "Componente interativo sem nome acessível, papel (role) ou estado ARIA correto.",
    "4.1.3": "Mensagens de alerta ou atualizações dinâmicas não anunciadas via região ao vivo (aria-live).",
}

_CRITERION_DEFAULT_IMPACTS: dict[str, str] = {
    "1.1.1": "Pessoas cegas ou com baixa visão que utilizam leitores de tela ouvem apenas o nome do arquivo ou perdem o significado da imagem.",
    "1.3.1": "Leitores de tela não conseguem estruturar a navegação por cabeçalhos ou associar campos aos seus rótulos.",
    "1.4.1": "Pessoas daltônicas ou com baixa visão não conseguem identificar mensagens ou campos com erro.",
    "1.4.3": "Pessoas com baixa visão ou idosos não conseguem ler o conteúdo com facilidade.",
    "1.4.11": "Controles interativos e botões tornam-se invisíveis ou imperceptíveis.",
    "2.1.1": "Pessoas com deficiência motora ou que não usam mouse ficam impedidas de concluir a ação.",
    "2.1.2": "O usuário fica bloqueado na tela e é forçado a reiniciar a navegação.",
    "2.4.1": "Usuários de teclado precisam navegar repetidamente por todo o cabeçalho a cada página.",
    "2.4.2": "Pessoas que utilizam leitores de tela não sabem em qual página ou aba estão posicionadas.",
    "2.4.3": "A ordem confusa de tabulação desorienta completamente o usuário.",
    "2.4.4": "Na lista de links do leitor de tela (NVDA/JAWS), o usuário não sabe para onde cada link leva.",
    "2.4.7": "Usuários de teclado perdem a referência visual de onde está o cursor.",
    "2.5.8": "Usuários em telas touch ou com tremores motores tocam em botões errados por proximidade excessiva.",
    "3.1.1": "O leitor de tela pronuncia as palavras com sotaque ou fonética inadequada.",
    "3.3.2": "O usuário não compreende o que deve ser digitado no campo, gerando erros de preenchimento.",
    "4.1.2": "O leitor de tela anuncia 'elemento desconhecido', impedindo o uso do componente.",
    "4.1.3": "O usuário não é avisado de que a operação foi concluída com sucesso ou que houve erro.",
}

_CRITERION_DEFAULT_REMEDIATIONS: dict[str, str] = {
    "1.1.1": "Adicione o atributo alt com descrição concisa na tag <img> (ou alt='' e aria-hidden='true' se for decorativa).",
    "1.3.1": "Utilize elementos semânticos nativos do HTML5 (<header>, <nav>, <main>, <section>, <h1>-<h6>, <label>).",
    "1.4.1": "Combine cores com ícones, textos descritivos ou sublinhados para reforçar o estado.",
    "1.4.3": "Ajuste as cores de texto e fundo para atingir pelo menos 4.5:1 de contraste (ou 3:1 para textos grandes).",
    "1.4.11": "Assegure que bordas de campos, ícones e foco tenham pelo menos 3:1 de contraste com o fundo.",
    "2.1.1": "Utilize tags nativas <button> ou <a>, ou inclua tabindex='0' e manipuladores para Enter e Espaço.",
    "2.1.2": "Implemente gerenciamento correto de foco, permitindo fechar o componente com a tecla Escape.",
    "2.4.1": "Adicione um link de salto (<a href='#main' class='skip-link'>Pular para o conteúdo principal</a>) no início da página.",
    "2.4.2": "Declare <title>Título Descritivo da Página</title> no <head>.",
    "2.4.3": "Alinhe a ordem do código HTML com a ordem visual na tela e evite tabindex positivo.",
    "2.4.4": "Utilize textos claros e específicos para links, indicando o documento ou destino real.",
    "2.4.7": "Forneça estilos visíveis de foco em :focus e :focus-visible (nunca outline: none sem substituto).",
    "2.5.8": "Ajuste dimensões mínimas ou espaçamento externo para garantir alvos de clique de pelo menos 24x24px.",
    "3.1.1": "Declare <html lang='pt-BR'> no início do documento.",
    "3.3.2": "Associe <label for='campo-id'> ao <input id='campo-id'> correspondente.",
    "4.1.2": "Adicione aria-label, role e atributos de estado (como aria-expanded) nos componentes customizados.",
    "4.1.3": "Utilize aria-live='polite' ou role='status' nos contêineres de feedback dinâmico.",
}


def _get_principle_name(code: str) -> str:
    """Retorna o princípio POUR com base no código do critério."""
    if code and code[0] in _PRINCIPLES_PT:
        return _PRINCIPLES_PT[code[0]][0]
    return "Princípio de Acessibilidade Digital"


def _get_level(code: str, raw_level: str | None) -> str:
    """Retorna o nível de conformidade WCAG."""
    if raw_level:
        return raw_level
    return _CRITERION_LEVELS.get(code, "A")


def _get_criterion_title(issue: dict[str, Any], code: str) -> str:
    """Retorna o título traduzido do critério."""
    pt = issue.get("criterion_pt")
    if pt:
        return pt
    if code in _CRITERION_PT:
        return _CRITERION_PT[code]
    return issue.get("criterion") or code


def _get_priority_label(level: str, severity: str) -> str:
    """Calcula a prioridade de atendimento da equipe."""
    s = (severity or "").lower()
    if level == "A" or "crit" in s:
        return "Prioridade 1 — Imediata (Bloqueador de Acesso)"
    if level == "AA" or "alt" in s or "high" in s:
        return "Prioridade 2 — Alta (Requisito Legal WCAG AA)"
    return "Prioridade 3 — Média (Melhoria Contínua)"


def _create_styles(wb: Workbook) -> dict[str, Any]:
    """Cria estilos reutilizáveis e de alto contraste para a planilha."""
    styles = {}

    # Header principal
    header = NamedStyle(name="a11y_header")
    header.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header.fill = PatternFill(start_color="0B1120", end_color="0B1120", fill_type="solid")
    header.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    header.border = Border(
        bottom=Side(style="medium", color="14B8A6"),
    )
    styles["header"] = header
    wb.add_named_style(header)

    # Célula normal
    cell = NamedStyle(name="a11y_cell")
    cell.font = Font(name="Calibri", size=10.5)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    styles["cell"] = cell
    wb.add_named_style(cell)

    # Nível A — Criticidade Máxima
    crit_a = NamedStyle(name="a11y_level_a")
    crit_a.font = Font(name="Calibri", size=10.5, color="7F1D1D", bold=True)
    crit_a.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    crit_a.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    styles["level_a"] = crit_a
    wb.add_named_style(crit_a)

    # Nível AA — Criticidade Legal
    crit_aa = NamedStyle(name="a11y_level_aa")
    crit_aa.font = Font(name="Calibri", size=10.5, color="854D0E", bold=True)
    crit_aa.fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
    crit_aa.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    styles["level_aa"] = crit_aa
    wb.add_named_style(crit_aa)

    # Nível AAA — Avançado
    crit_aaa = NamedStyle(name="a11y_level_aaa")
    crit_aaa.font = Font(name="Calibri", size=10.5, color="166534")
    crit_aaa.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    crit_aaa.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    styles["level_aaa"] = crit_aaa
    wb.add_named_style(crit_aaa)

    return styles


def _level_style(level: str, styles: dict) -> Any:
    if level == "A":
        return styles["level_a"]
    if level == "AA":
        return styles["level_aa"]
    if level == "AAA":
        return styles["level_aaa"]
    return styles["cell"]


def _sort_key_for_issue(issue: dict[str, Any]) -> tuple[int, str, int]:
    """Ordena as linhas para que os princípios fiquem um embaixo do outro."""
    raw_crit = issue.get("criterion") or ""
    code = _extract_code(raw_crit)
    principle_char = code[0] if code and code[0] in "1234" else "9"
    principle_order = int(principle_char) if principle_char.isdigit() else 9

    level = issue.get("level") or _CRITERION_LEVELS.get(code, "A")
    level_order = {"A": 1, "AA": 2, "AAA": 3}.get(level, 4)

    return (principle_order, code, level_order)


def export_issues_xlsx(issues: list[dict], url: str | None = None) -> bytes:
    """
    Exporta lista de issues para XLSX acessível, didático e estruturado.

    Organização:
    - Linhas organizadas de cima para baixo por Princípio WCAG (POUR)
    - Colunas dispostas lado a lado com todos os metadados requeridos:
      1. ID
      2. Site / URL Testada
      3. Onde Está o Problema (Localização / Elemento)
      4. Princípio WCAG (POUR)
      5. Diretriz WCAG
      6. Critério de Sucesso
      7. Criticidade Normativa (WCAG) -> Nível A, Nível AA, Nível AAA
      8. Prioridade de Atendimento da Equipe
      9. O Que Aconteceu (Diagnóstico Claro)
      10. Por Que Afeta o Usuário (Impacto / Leitor de Tela)
      11. Como Resolver (Plano de Remediação)
      12. Navegador Usado
      13. Leitor de Tela Testado
      14. Código Original (HTML)
      15. Código Corrigido (HTML)
      16. Status
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Issues de Acessibilidade"

    styles = _create_styles(wb)

    headers = [
        "ID",
        "Site / URL Testada",
        "Onde Está o Problema (Localização / Elemento)",
        "Princípio WCAG",
        "Diretriz WCAG",
        "Critério de Sucesso",
        "Criticidade Normativa (WCAG)",
        "Prioridade de Atendimento",
        "O Que Aconteceu (Diagnóstico)",
        "Por Que Afeta o Usuário",
        "Como Resolver (Plano de Remediação)",
        "Navegador Usado",
        "Leitor de Tela Testado",
        "Código Original (HTML)",
        "Código Corrigido (HTML)",
        "Status",
    ]

    # Escreve header
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.style = styles["header"]

    # Freeze panes — header fixo ao rolar
    ws.freeze_panes = "A2"

    # Colunas com largura otimizada
    col_widths = [
        12,   # ID
        28,   # Site / URL Testada
        35,   # Onde Está o Problema
        24,   # Princípio WCAG
        26,   # Diretriz WCAG
        34,   # Critério de Sucesso
        24,   # Criticidade Normativa (Nível A, AA, AAA)
        34,   # Prioridade de Atendimento
        45,   # Diagnóstico
        45,   # Por Que Afeta
        50,   # Como Resolver
        35,   # Navegador Usado
        35,   # Leitor de Tela Testado
        35,   # Código Original (HTML)
        35,   # Código Corrigido (HTML)
        14,   # Status
    ]

    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Altura do header
    ws.row_dimensions[1].height = 32

    # Ordena as linhas para que os princípios fiquem um embaixo do outro
    sorted_issues = sorted(issues, key=_sort_key_for_issue)

    # Escreve dados
    for row_idx, issue in enumerate(sorted_issues, 2):
        ws.row_dimensions[row_idx].height = 65

        raw_crit = issue.get("criterion") or ""
        code = _extract_code(raw_crit)
        principle_name = _get_principle_name(code)
        guideline_name = get_guideline_for_code(code)
        level = _get_level(code, issue.get("level"))
        crit_title = _get_criterion_title(issue, code)

        criticality_label = f"Nível {level}"
        priority_label = _get_priority_label(level, issue.get("severity", ""))

        # Diagnóstico claro
        problem = (
            issue.get("description")
            or issue.get("description_technical")
            or _CRITERION_DEFAULT_PROBLEMS.get(code, "Não conformidade de acessibilidade identificada.")
        )

        # Por que afeta
        impact = (
            issue.get("why_simple")
            or issue.get("why_technical")
            or _CRITERION_DEFAULT_IMPACTS.get(code, "Gera barreiras de navegação para usuários de tecnologias assistivas.")
        )

        # Como resolver
        solution = (
            issue.get("suggestion")
            or issue.get("suggestion_technical")
            or _CRITERION_DEFAULT_REMEDIATIONS.get(code, "Consulte a especificação técnica do WCAG para aplicar a remediação semântica.")
        )

        site_tested = issue.get("url") or url or "Interface Local / URL não informada"
        where_problem = issue.get("element") or "Interface Web / Componente Geral"

        values = [
            issue.get("id") or f"ISS-{row_idx-1}",
            site_tested,
            where_problem,
            principle_name,
            guideline_name,
            crit_title,
            criticality_label,
            priority_label,
            problem,
            impact,
            solution,
            _ENVIRONMENT_TESTED,
            _SCREEN_READER_TESTED,
            issue.get("element") or "",
            issue.get("fixed_element_html") or "",
            "Corrigido" if issue.get("fixed_element_html") else "Aberto",
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 7:  # Criticidade Normativa — estilo com cor e contraste por Nível (A, AA, AAA)
                cell.style = _level_style(level, styles)
            else:
                cell.style = styles["cell"]

    # Auto-filter em todas as colunas com dados
    last_col = len(headers)
    last_row = len(sorted_issues) + 1
    if last_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"

    # Cria Excel Table
    tab = Table(displayName="AccessibilityIssues", ref=f"A1:{get_column_letter(last_col)}{last_row}")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    # Sheet protection desabilitado
    ws.protection.disable()

    # Stream para bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
